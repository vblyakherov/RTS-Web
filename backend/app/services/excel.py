import io
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

import xlsxwriter
from openpyxl import load_workbook

from app.core.columns import (
    SITE_COLUMNS,
    SYNC_KEY_COLUMN,
    get_column_by_db_name,
    get_column_by_header,
    normalize_excel_header,
)
from app.models.site import Site
from app.services.ucn_template import apply_template_derivations


EXPORT_COLUMNS = [(SYNC_KEY_COLUMN.db_name, SYNC_KEY_COLUMN.excel_header)]
EXPORT_COLUMNS.extend((c.db_name, c.excel_header) for c in SITE_COLUMNS)

TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "templates" / "sync_template.xlsm"
DATA_SHEET_NAME = "Data"
SYNC_SHEET_NAME = "Sync"
VBA_PROJECT_PART = "xl/vbaProject.bin"
DATA_SHEET_PROTECTION_PASSWORD = "RTKS_SYNC_DATA"
XML_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
}

REQUIRED_IMPORT_HEADERS = {
    normalize_excel_header(SYNC_KEY_COLUMN.excel_header),
    "Регион",
    "Наименование НП",
}


class ExcelTemplateError(RuntimeError):
    """Ошибка при работе с XLSM-шаблоном экспорта."""


@dataclass(frozen=True)
class TemplateMetadata:
    vba_project: bytes
    workbook_codename: str | None
    data_sheet_codename: str | None


def export_sites_to_excel(
    sites: list[Site],
    auth_token: str | None = None,
    username: str | None = None,
    project_id: int | None = None,
) -> bytes:
    template = _load_template_metadata()
    baseline = _get_export_last_sync_at(sites)

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    try:
        _attach_vba_project(workbook, template)
        _build_data_sheet(workbook, sites, data_sheet_codename=template.data_sheet_codename)
        _build_sync_sheet(workbook)
        _build_auxiliary_sheets(
            workbook,
            baseline=baseline,
            auth_token=auth_token,
            username=username,
            project_id=project_id,
        )
    finally:
        workbook.close()

    output.seek(0)
    return output.getvalue()


def parse_excel_import(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    errors: list[str] = []
    rows: list[dict[str, Any]] = []

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        return [], [f"Не удалось прочитать файл: {exc}"]

    worksheet = workbook[workbook.sheetnames[0]]
    header_by_index = {
        idx: normalize_excel_header(cell.value)
        for idx, cell in enumerate(worksheet[1], start=1)
        if normalize_excel_header(cell.value)
    }

    missing = REQUIRED_IMPORT_HEADERS - set(header_by_index.values())
    if missing:
        return [], [f"Отсутствуют обязательные колонки: {', '.join(sorted(missing))}"]

    key_header = normalize_excel_header(SYNC_KEY_COLUMN.excel_header)

    for row_idx in range(2, worksheet.max_row + 1):
        if not any(worksheet.cell(row_idx, col_idx).value not in (None, "") for col_idx in header_by_index):
            continue

        parsed: dict[str, Any] = {}
        site_id: str | None = None
        row_errors: list[str] = []

        for col_idx, header in header_by_index.items():
            raw_value = worksheet.cell(row_idx, col_idx).value
            if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
                continue

            if header == key_header:
                site_id = str(raw_value).strip().upper()
                continue

            col_def = get_column_by_header(header)
            if col_def is None:
                continue

            try:
                parsed[col_def.db_name] = _coerce_excel_value(raw_value, col_def.python_type)
            except ValueError as exc:
                row_errors.append(f"поле '{header}': {exc}")

        if not site_id:
            errors.append(f"Строка {row_idx}: пустой ID объекта — пропущена")
            continue

        if row_errors:
            for err in row_errors:
                errors.append(f"Строка {row_idx} ({site_id}): {err}")

        parsed["site_id"] = site_id
        parsed = apply_template_derivations(parsed)

        if not parsed.get("name"):
            errors.append(f"Строка {row_idx} ({site_id}): пустое наименование НП — пропущена")
            continue
        if not parsed.get("region"):
            errors.append(f"Строка {row_idx} ({site_id}): пустой регион — пропущена")
            continue

        rows.append(parsed)

    return rows, errors


def _site_to_export_row(site: Site) -> list[Any]:
    row: list[Any] = []
    for db_name, _ in EXPORT_COLUMNS:
        value = getattr(site, db_name, None)
        if isinstance(value, datetime):
            row.append(_fmt_dt(value))
        else:
            row.append(value if value is not None else "")
    return row


def _load_template_metadata() -> TemplateMetadata:
    if not TEMPLATE_PATH.exists():
        raise ExcelTemplateError(
            "Не найден шаблон XLSM: backend/templates/sync_template.xlsm. "
            "Создайте его в Excel по инструкции из backend/templates/README.md."
        )

    try:
        with ZipFile(TEMPLATE_PATH) as archive:
            vba_project = _read_vba_project(archive)
            workbook_codename, data_sheet_codename = _read_template_codenames(archive)
    except (BadZipFile, KeyError, ElementTree.ParseError) as exc:
        raise ExcelTemplateError(
            "Не удалось прочитать VBA-часть шаблона backend/templates/sync_template.xlsm."
        ) from exc

    return TemplateMetadata(
        vba_project=vba_project,
        workbook_codename=workbook_codename,
        data_sheet_codename=data_sheet_codename,
    )


def _read_vba_project(archive: ZipFile) -> bytes:
    try:
        return archive.read(VBA_PROJECT_PART)
    except KeyError as exc:
        raise ExcelTemplateError(
            f"В шаблоне {TEMPLATE_PATH.name} отсутствует {VBA_PROJECT_PART}."
        ) from exc


def _read_template_codenames(archive: ZipFile) -> tuple[str | None, str | None]:
    workbook_xml = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    workbook_pr = workbook_xml.find("main:workbookPr", XML_NS)
    workbook_codename = workbook_pr.get("codeName") if workbook_pr is not None else None

    data_sheet = None
    for sheet in workbook_xml.findall("main:sheets/main:sheet", XML_NS):
        if sheet.get("name") == DATA_SHEET_NAME:
            data_sheet = sheet
            break

    if data_sheet is None:
        raise ExcelTemplateError(
            f"В шаблоне {TEMPLATE_PATH.name} отсутствует лист '{DATA_SHEET_NAME}'."
        )

    rel_id = data_sheet.get(f"{{{XML_NS['rel']}}}id")
    if not rel_id:
        raise ExcelTemplateError(
            f"Не удалось определить XML-связь для листа '{DATA_SHEET_NAME}' в {TEMPLATE_PATH.name}."
        )

    rels_xml = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    sheet_target = None
    for rel in rels_xml.findall("pkg:Relationship", XML_NS):
        if rel.get("Id") == rel_id:
            sheet_target = rel.get("Target")
            break

    if not sheet_target:
        raise ExcelTemplateError(
            f"Не удалось найти XML-файл листа '{DATA_SHEET_NAME}' в {TEMPLATE_PATH.name}."
        )

    sheet_xml_path = "xl/" + sheet_target.lstrip("/")
    sheet_xml = ElementTree.fromstring(archive.read(sheet_xml_path))
    sheet_pr = sheet_xml.find("main:sheetPr", XML_NS)
    sheet_codename = sheet_pr.get("codeName") if sheet_pr is not None else None
    return workbook_codename, sheet_codename


def _attach_vba_project(workbook: xlsxwriter.Workbook, template: TemplateMetadata) -> None:
    workbook.add_vba_project(io.BytesIO(template.vba_project), is_stream=True)
    if template.workbook_codename:
        workbook.set_vba_name(template.workbook_codename)


def _build_sync_sheet(workbook: xlsxwriter.Workbook) -> None:
    sync_ws = workbook.add_worksheet(SYNC_SHEET_NAME)
    sync_ws.activate()
    sync_ws.set_first_sheet()
    sync_ws.hide_gridlines(2)
    sync_ws.set_default_row(22)
    sync_ws.set_column("A:A", 3)
    sync_ws.set_column("B:B", 18)
    sync_ws.set_column("C:C", 42)
    sync_ws.set_column("D:D", 20)

    title_fmt = workbook.add_format({
        "bold": True,
        "font_size": 20,
        "font_color": "#0F172A",
    })
    text_fmt = workbook.add_format({
        "font_size": 11,
        "font_color": "#334155",
        "text_wrap": True,
        "valign": "top",
    })
    link_fmt = workbook.add_format({
        "font_size": 11,
        "font_color": "#2563EB",
        "underline": 1,
    })

    sync_ws.merge_range("B2:D2", "Excel Sync", title_fmt)
    sync_ws.write("B4", "1.", text_fmt)
    sync_ws.write("C4", "Разрешите макросы при открытии файла.", text_fmt)
    sync_ws.write("B5", "2.", text_fmt)
    sync_ws.write("C5", "Редактируйте существующие объекты на листе Data.", text_fmt)
    sync_ws.write("B6", "3.", text_fmt)
    sync_ws.write("C6", "Добавление и удаление строк заблокировано: меняйте только значения ячеек.", text_fmt)
    sync_ws.write("B7", "4.", text_fmt)
    sync_ws.write("C7", "Возвращайтесь сюда и нажимайте кнопку синхронизации.", text_fmt)
    sync_ws.write_url("C8", f"internal:{DATA_SHEET_NAME}!A1", link_fmt, string="Открыть лист Data")
    sync_ws.insert_button("B10", {
        "macro": "SyncNow",
        "caption": "Синхронизировать",
        "width": 220,
        "height": 38,
        "x_offset": 6,
        "y_offset": 4,
    })


def _build_data_sheet(
    workbook: xlsxwriter.Workbook,
    sites: list[Site],
    data_sheet_codename: str | None,
) -> None:
    ws = workbook.add_worksheet(DATA_SHEET_NAME)
    if data_sheet_codename:
        ws.set_vba_name(data_sheet_codename)

    header_fmt = workbook.add_format({
        "bold": True,
        "font_color": "#FFFFFF",
        "bg_color": "#2563EB",
        "align": "center",
        "valign": "vcenter",
        "border": 1,
    })
    editable_fmt = workbook.add_format({"locked": False})

    headers = [col_label for _, col_label in EXPORT_COLUMNS]
    for col_idx, header in enumerate(headers):
        ws.write(0, col_idx, header, header_fmt)

    for row_idx, site in enumerate(sites, start=1):
        row_data = _site_to_export_row(site)
        for col_idx, value in enumerate(row_data):
            ws.write(row_idx, col_idx, value)

    for col_idx, (db_name, _) in enumerate(EXPORT_COLUMNS):
        col_def = get_column_by_db_name(db_name)
        width = col_def.excel_width if col_def else 18
        ws.set_column(col_idx, col_idx, width, editable_fmt)

    ws.freeze_panes(1, 0)
    ws.autofilter(0, 0, max(len(sites), 1), len(EXPORT_COLUMNS) - 1)
    ws.protect(DATA_SHEET_PROTECTION_PASSWORD, {
        "autofilter": True,
        "sort": True,
        "select_locked_cells": False,
        "select_unlocked_cells": True,
    })


def _build_auxiliary_sheets(
    workbook: xlsxwriter.Workbook,
    baseline: str,
    auth_token: str | None,
    username: str | None,
    project_id: int | None,
) -> None:
    config_ws = workbook.add_worksheet("_Config")
    config_ws.very_hidden()
    config_ws.write_row(0, 0, ("last_sync_at", baseline))
    config_ws.write_row(1, 0, ("auth_token", auth_token or ""))
    config_ws.write_row(2, 0, ("username", username or ""))
    config_ws.write_row(3, 0, ("project_id", project_id or ""))

    dirty_ws = workbook.add_worksheet("_DirtyTracker")
    dirty_ws.very_hidden()


def _get_export_last_sync_at(sites: list[Site]) -> str:
    updated_values = []
    for site in sites:
        updated_at = getattr(site, "updated_at", None)
        if updated_at is None:
            continue
        if updated_at.tzinfo is None:
            updated_at = updated_at.replace(tzinfo=timezone.utc)
        updated_values.append(updated_at)

    if not updated_values:
        return datetime.now(timezone.utc).isoformat()

    return max(updated_values).isoformat()


def _fmt_dt(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return dt.strftime("%d.%m.%Y")


def _coerce_excel_value(value: Any, target_type: type) -> Any:
    if value is None:
        return None

    if target_type == datetime:
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())

        raw = str(value).strip()
        if not raw:
            return None

        for fmt in (
            "%d.%m.%Y",
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d.%m.%Y %H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
        ):
            try:
                return datetime.strptime(raw, fmt)
            except ValueError:
                continue
        raise ValueError(f"не удалось распознать дату '{value}'")

    if target_type == float:
        return float(value)

    if target_type == int:
        return int(float(value))

    normalized = str(value).strip()
    return normalized or None
